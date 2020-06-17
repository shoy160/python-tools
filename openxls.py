from openpyxl import Workbook, load_workbook
from business.shop import ShopService


def read_xlsx(path):
    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    rows = ws.rows
    cols = ws.columns
    lines = []
    for row in rows:
        line = [col.value for col in row]
        lines.append(line)
    wb.close()
    return lines


def shop_sync():
    service = ShopService()
    # path = "doc/经销商机构最终版.xlsx"
    path = "doc/unknown.xlsx"
    lines = read_xlsx(path)
    unknow_lines = []
    index = 0
    for line in lines[1:]:
        index += 1
        print('\033[1;35m row %d:%s \033[0m' % (index, line[3]))
        service.check_shop(line, lambda x: unknow_lines.append(x))
    # 保存未知的数据
    wb = Workbook()
    try:
        ws = wb.create_sheet()
        # head
        ws.append(lines[0])
        for line in unknow_lines:
            ws.append(line)
    finally:
        wb.save('doc/unknown.xlsx')
        wb.close()


def order_compare():
    order_path = 'doc/order_190418.xlsx'
    sign_path = 'doc/sign_190418.xlsx'

    order_lines = read_xlsx(order_path)
    sign_lines = read_xlsx(sign_path)
    policy_list = []
    for s_line in sign_lines[1:]:
        policy = s_line[2]
        exists = False
        for o_line in order_lines[1:]:
            if str(o_line[1]) == str(policy):
                exists = True
        if not exists:
            print(policy)
            # policy_list.append(policy)

    # print(','.join(policy_list))
